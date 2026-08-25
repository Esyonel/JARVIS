PLUGIN = {
    "name": "excel_automation",
    "description": "Automates processing of large Excel workbooks using openpyxl/pandas. Allows applying formulas, simple data transformations and saving the result. Ideal for backend utilities and workflow automation.",
    "parameters": {
        "type": "OBJECT",
        "properties": {
            "input_path": {
                "type": "string",
                "description": "Absolute or relative path to the source Excel file (*.xlsx)."
            },
            "output_path": {
                "type": "string",
                "description": "Path where the modified workbook will be saved. If omitted, the original file is overwritten."
            },
            "formulas": {
                "type": "array",
                "description": "List of formula specifications to apply.",
                "items": {
                    "type": "object",
                    "properties": {
                        "sheet": {"type": "string", "description": "Name of the worksheet where the formula will be placed."},
                        "cell": {"type": "string", "description": "A1‑style cell reference (e.g., B2)."},
                        "formula": {"type": "string", "description": "Excel formula string without the leading '=' (e.g., 'SUM(A1:A10)')."}
                    },
                    "required": ["sheet", "cell", "formula"]
                }
            },
            "pandas_operations": {
                "type": "array",
                "description": "Optional pandas based transformations. Each entry is a Python expression evaluated with a temporary DataFrame variable `df`. The expression must return a DataFrame which will replace the original sheet.",
                "items": {
                    "type": "object",
                    "properties": {
                        "sheet": {"type": "string", "description": "Worksheet name to load into pandas."},
                        "operation": {"type": "string", "description": "Python code string executed with `df` in the local namespace. Example: 'df[\"total\"] = df[\"price\"] * df[\"qty\"]'"}
                    },
                    "required": ["sheet", "operation"]
                }
            }
        },
        "required": ["input_path"]
    }
}

def run(parameters: dict, player=None, session_memory=None) -> str:
    """Execute the Excel automation plugin.

    The function never raises; any exception is caught and turned into a short
    spoken message. It returns a concise status string that JARVIS can speak.
    """
    import os
    try:
        input_path = parameters.get("input_path")
        if not input_path:
            return "Input path is missing. Please provide a valid Excel file path."
        if not os.path.isfile(input_path):
            return f"The file {input_path} does not exist."

        # Determine output location
        output_path = parameters.get("output_path") or input_path

        # Import heavy libraries lazily – they are optional dependencies.
        try:
            from openpyxl import load_workbook
        except Exception as e:
            return f"openpyxl is required for Excel automation but could not be loaded: {e}"

        # Load workbook with read/write mode (keep existing formulas).
        wb = load_workbook(filename=input_path, data_only=False)

        # Apply simple formula updates if provided.
        formulas = parameters.get("formulas", [])
        for spec in formulas:
            sheet_name = spec.get("sheet")
            cell = spec.get("cell")
            formula = spec.get("formula")
            if not (sheet_name and cell and formula):
                continue  # skip incomplete entries silently
            if sheet_name not in wb.sheetnames:
                return f"Sheet {sheet_name} not found in the workbook."
            ws = wb[sheet_name]
            ws[cell].value = f"={formula}"

        # Optional pandas based transformations.
        pandas_ops = parameters.get("pandas_operations", [])
        if pandas_ops:
            try:
                import pandas as pd
            except Exception as e:
                return f"pandas is required for data‑frame operations but could not be loaded: {e}"
            for op in pandas_ops:
                sheet_name = op.get("sheet")
                operation_code = op.get("operation")
                if not (sheet_name and operation_code):
                    continue
                if sheet_name not in wb.sheetnames:
                    return f"Sheet {sheet_name} not found for pandas operation."
                # Read sheet into DataFrame
                df = pd.read_excel(input_path, sheet_name=sheet_name, engine='openpyxl')
                # Execute user‑provided operation safely.
                local_ns = {"df": df}
                try:
                    exec(operation_code, {}, local_ns)
                except Exception as e:
                    return f"Error executing pandas operation on sheet {sheet_name}: {e}"
                # Expect the operation to modify `df` in place or replace it.
                df_result = local_ns.get("df")
                if df_result is None:
                    return f"Pandas operation on sheet {sheet_name} did not produce a DataFrame."
                # Write back to workbook using openpyxl's writer.
                # First clear existing rows.
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row)
                for r_idx, row in enumerate(df_result.itertuples(index=False, name=None), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                # Write header
                for c_idx, col_name in enumerate(df_result.columns, start=1):
                    ws.cell(row=1, column=c_idx, value=col_name)

        # Save workbook
        wb.save(output_path)
        if output_path == input_path:
            return "Excel file has been updated successfully."
        else:
            return f"Excel file processed and saved to {output_path}."
    except Exception as e:
        return f"An unexpected error occurred during Excel automation: {e}"