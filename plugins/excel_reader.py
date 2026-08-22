"""
JARVIS plugin — Excel reader/analyst (existing .xlsx → spoken answer).

Complements excel_writer.py (which only *creates* new spreadsheets). This
plugin opens an EXISTING .xlsx/.xlsm/.csv and answers questions about it:
totals, max/min, row counts, or just "what's in this file". Small files are
summarized in full; large ones are aggregated (sums/counts per numeric
column) so the answer stays short enough to speak.
"""

import csv
from pathlib import Path

PLUGIN = {
    "name": "excel_reader",
    "description": (
        "Opens an EXISTING Excel (.xlsx/.xlsm) or .csv file and answers "
        "questions about its contents — totals, max/min, row/column counts, "
        "or a general summary. Use for: 'bu tablodaki toplam ne kadar', "
        "'şu dosyada kaç satır var', 'en yüksek değer hangi satırda', "
        "'masaüstündeki filoyu.xlsx dosyasını özetle'. Use excel_writer "
        "instead when the user wants a NEW file created, not an existing one read."
    ),
    "parameters": {
        "type": "OBJECT",
        "properties": {
            "path": {
                "type": "STRING",
                "description": "Full or partial path/filename of the spreadsheet to read.",
            },
            "question": {
                "type": "STRING",
                "description": (
                    "What the user wants to know about the file, verbatim "
                    "(e.g. 'toplam satış ne kadar', 'özetle'). Optional — "
                    "if omitted, a general summary is returned."
                ),
            },
            "sheet": {
                "type": "STRING",
                "description": "Sheet name, if the workbook has more than one and a specific one is wanted.",
            },
        },
        "required": ["path"],
    },
}

_SEARCH_DIRS = [Path.home() / "Desktop", Path.home() / "Documents", Path.home() / "Downloads"]
_MAX_ROWS_FULL = 30


def run(parameters: dict, player=None, session_memory=None) -> str:
    raw_path = (parameters.get("path") or "").strip()
    question = (parameters.get("question") or "").strip().lower()
    sheet_name = (parameters.get("sheet") or "").strip() or None

    if not raw_path:
        msg = "Sir, I need a file name or path to read."
        _log(msg, player)
        return msg

    path = _resolve_path(raw_path)
    if path is None:
        msg = f"Sir, I couldn't find a file matching '{raw_path}' on the Desktop, Documents, or Downloads."
        _log(msg, player)
        return msg

    try:
        if path.suffix.lower() == ".csv":
            headers, rows = _read_csv(path)
        else:
            headers, rows = _read_xlsx(path, sheet_name)
    except Exception as e:
        msg = f"Sir, I couldn't open '{path.name}': {e}"
        _log(msg, player)
        return msg

    if not headers and not rows:
        msg = f"'{path.name}' appears to be empty."
        _log(msg, player)
        return msg

    result = _answer(path.name, headers, rows, question)
    _log(result, player)
    return result


def _resolve_path(raw: str) -> Path | None:
    p = Path(raw).expanduser()
    if p.exists() and p.is_file():
        return p
    name_lower = raw.lower()
    for d in _SEARCH_DIRS:
        if not d.exists():
            continue
        for f in d.glob("*"):
            if f.is_file() and f.suffix.lower() in (".xlsx", ".xlsm", ".csv"):
                if name_lower in f.name.lower():
                    return f
    return None


def _read_csv(path: Path) -> tuple[list[str], list[list]]:
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.reader(f)
        data = list(reader)
    if not data:
        return [], []
    return data[0], data[1:]


def _read_xlsx(path: Path, sheet_name: str | None) -> tuple[list[str], list[list]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return [], []
        rows = [list(r) for r in rows_iter]
        return headers, rows
    finally:
        wb.close()


def _answer(filename: str, headers: list[str], rows: list[list], question: str) -> str:
    n_rows, n_cols = len(rows), len(headers)

    if "kaç sat" in question or "row count" in question or "how many row" in question:
        return f"'{filename}' has {n_rows} data rows and {n_cols} columns."

    numeric_cols = _numeric_columns(headers, rows)

    wants_total = any(w in question for w in ("toplam", "total", "sum"))
    wants_max = any(w in question for w in ("en yüksek", "en büyük", "max", "highest"))
    wants_min = any(w in question for w in ("en düşük", "en küçük", "min", "lowest"))

    if (wants_total or wants_max or wants_min) and numeric_cols:
        parts = []
        for col_name, values in numeric_cols:
            if wants_total:
                parts.append(f"{col_name} toplamı: {sum(values):,.2f}")
            if wants_max:
                parts.append(f"{col_name} en yüksek: {max(values):,.2f}")
            if wants_min:
                parts.append(f"{col_name} en düşük: {min(values):,.2f}")
        if parts:
            return f"'{filename}' ({n_rows} satır) — " + " | ".join(parts)

    if n_rows <= _MAX_ROWS_FULL:
        preview = "; ".join(
            ", ".join(f"{h}={v}" for h, v in zip(headers, row) if v not in (None, ""))
            for row in rows
        )
        return f"'{filename}' — sütunlar: {', '.join(headers)}. İçerik: {preview}"[:3000]

    summary_bits = [f"'{filename}': {n_rows} satır, {n_cols} sütun ({', '.join(headers)})."]
    for col_name, values in numeric_cols[:5]:
        summary_bits.append(
            f"{col_name}: toplam {sum(values):,.2f}, ortalama {sum(values)/len(values):,.2f}, "
            f"min {min(values):,.2f}, max {max(values):,.2f}"
        )
    return " ".join(summary_bits)[:3000]


def _numeric_columns(headers: list[str], rows: list[list]) -> list[tuple[str, list[float]]]:
    out = []
    for i, name in enumerate(headers):
        values = []
        for row in rows:
            if i >= len(row):
                continue
            v = row[i]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                values.append(float(v))
        if len(values) >= max(1, len(rows) // 2):
            out.append((name or f"col{i+1}", values))
    return out


def _log(message: str, player=None) -> None:
    print(f"[ExcelReader] {message[:200]}")
    if player:
        try:
            player.write_log(f"JARVIS: {message}")
        except Exception:
            pass
