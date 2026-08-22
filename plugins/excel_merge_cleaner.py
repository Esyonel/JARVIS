"""
JARVIS plugin — merge & clean multiple Excel/CSV files into one.

Finds files by name fragment on the Desktop/Documents/Downloads, stacks rows
from every matching file that shares the same header row, drops exact
duplicate rows, and writes a single cleaned .xlsx to the Desktop. Useful for
the "same report, five backup copies" pattern (e.g. "rapor.xlsx",
"rapor (yedek).xlsx", "rapor-son.xlsx").
"""

import csv
from datetime import datetime
from pathlib import Path

PLUGIN = {
    "name": "excel_merge_cleaner",
    "description": (
        "Merges several Excel/CSV files that share the same columns into one "
        "cleaned .xlsx file, removing exact duplicate rows, and saves the "
        "result to the Desktop. Use for: 'şu dosyaları birleştir', 'yedek "
        "kopyaları tek tabloda topla', 'bu tabloları birleştirip tekrarları "
        "temizle'. Give a name fragment shared by the files (e.g. 'mazot' "
        "matches every file with 'mazot' in its name)."
    ),
    "parameters": {
        "type": "OBJECT",
        "properties": {
            "name_fragment": {
                "type": "STRING",
                "description": "Text that appears in the names of all files to merge (e.g. 'mazot', 'rapor').",
            },
            "output_name": {
                "type": "STRING",
                "description": "Optional name for the merged output file (without extension).",
            },
        },
        "required": ["name_fragment"],
    },
}

_SEARCH_DIRS = [Path.home() / "Desktop", Path.home() / "Documents", Path.home() / "Downloads"]


def run(parameters: dict, player=None, session_memory=None) -> str:
    fragment = (parameters.get("name_fragment") or "").strip().lower()
    output_name = (parameters.get("output_name") or "").strip()

    if not fragment:
        msg = "Sir, I need a name fragment to find the files to merge."
        _log(msg, player)
        return msg

    matches = _find_matches(fragment)
    if len(matches) < 2:
        msg = (
            f"Sir, I only found {len(matches)} file(s) matching '{fragment}' — "
            "need at least two to merge."
        )
        _log(msg, player)
        return msg

    try:
        headers, all_rows, per_file, skipped = _collect_rows(matches)
        if headers is None:
            msg = f"Sir, none of the {len(matches)} matching files could be read."
            _log(msg, player)
            return msg

        unique_rows = _dedupe(all_rows)
        out_path = _write_output(headers, unique_rows, output_name)
    except Exception as e:
        msg = f"Sir, the merge failed: {e}"
        _log(msg, player)
        return msg

    msg = (
        f"Merged {per_file} file(s) matching '{fragment}' "
        f"({len(all_rows)} rows → {len(unique_rows)} after removing duplicates"
        f"{f', {skipped} file(s) skipped (different columns)' if skipped else ''}). "
        f"Saved to {out_path}."
    )
    _log(msg, player)
    return msg


def _find_matches(fragment: str) -> list[Path]:
    found = []
    for d in _SEARCH_DIRS:
        if not d.exists():
            continue
        for f in d.glob("*"):
            if f.is_file() and f.suffix.lower() in (".xlsx", ".xlsm", ".csv") and fragment in f.name.lower():
                found.append(f)
    return found


def _read_any(path: Path) -> tuple[list[str], list[list]]:
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            data = list(csv.reader(f))
        return (data[0], data[1:]) if data else ([], [])

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return [], []
        return headers, [list(r) for r in rows_iter]
    finally:
        wb.close()


def _collect_rows(paths: list[Path]):
    base_headers = None
    all_rows: list[tuple] = []
    used = 0
    skipped = 0

    for p in paths:
        try:
            headers, rows = _read_any(p)
        except Exception:
            skipped += 1
            continue
        if not headers:
            skipped += 1
            continue
        if base_headers is None:
            base_headers = headers
        elif [h.strip().lower() for h in headers] != [h.strip().lower() for h in base_headers]:
            skipped += 1
            continue
        all_rows.extend(tuple(r) for r in rows)
        used += 1

    return base_headers, all_rows, used, skipped


def _dedupe(rows: list[tuple]) -> list[tuple]:
    seen = set()
    out = []
    for r in rows:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def _write_output(headers: list[str], rows: list[tuple], output_name: str) -> Path:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))

    name = output_name or f"merged_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"

    desktop = Path.home() / "Desktop"
    out_dir = desktop if desktop.exists() else Path.home()
    out_path = out_dir / name
    wb.save(out_path)
    return out_path


def _log(message: str, player=None) -> None:
    print(f"[ExcelMergeCleaner] {message[:200]}")
    if player:
        try:
            player.write_log(f"JARVIS: {message}")
        except Exception:
            pass
