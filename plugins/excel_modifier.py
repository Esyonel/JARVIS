import os
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# Plugin metadata as required by the JARVIS framework
PLUGIN = {
    "name": "excel_modifier",
    "description": "Açık olan Excel dosyasını doğrudan değiştirebilir. Belirtilen sayfada, hücreye yeni bir değer yazar ve dosyayı kaydeder.",
    "parameters": {
        "type": "OBJECT",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Tam yoluyla Excel dosyasının yolu (örnek: C:/Users/Me/Documents/data.xlsx)."
            },
            "sheet_name": {
                "type": "string",
                "description": "Değişiklik yapılacak sayfanın adı."
            },
            "cell": {
                "type": "string",
                "description": "Excel hücre referansı (örnek: A1, B3)."
            },
            "value": {
                "type": "string",
                "description": "Hücreye yazılacak yeni değer."
            }
        },
        "required": ["file_path", "sheet_name", "cell", "value"]
    }
}

def _validate_parameters(params: dict) -> tuple[bool, str]:
    """Simple validation of required keys and file existence.
    Returns (True, "") if ok, otherwise (False, error_message).
    """
    for key in PLUGIN["parameters"]["required"]:
        if key not in params:
            return False, f"'{key}' parametresi eksik."
    path = params.get("file_path", "")
    if not isinstance(path, str) or not path:
        return False, "'file_path' geçerli bir string olmalı."
    if not os.path.isfile(path):
        return False, f"'{path}' dosyası bulunamadı."
    if load_workbook is None:
        return False, "openpyxl kütüphanesi yüklü değil."
    return True, ""

def run(parameters: dict, player: Any = None, session_memory: Any = None) -> str:
    """Modify a cell in an open Excel file.

    Args:
        parameters: Dictionary matching PLUGIN['parameters'] schema.
        player: (optional) JARVIS ses oynatıcı, kullanılmaz.
        session_memory: (optional) Oturum hafızası, kullanılmaz.

    Returns:
        Short Turkish string that JARVIS sesli olarak söyleyecek.
    """
    try:
        ok, msg = _validate_parameters(parameters)
        if not ok:
            return f"Hata: {msg}"

        file_path = parameters["file_path"]
        sheet_name = parameters["sheet_name"]
        cell = parameters["cell"]
        value = parameters["value"]

        # Load workbook with keep_vba=True to preserve macros if present
        wb = load_workbook(filename=file_path, keep_vba=True)
        if sheet_name not in wb.sheetnames:
            return f"Hata: '{sheet_name}' isimli sayfa bulunamadı."
        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(file_path)
        return f"{sheet_name} sayfasındaki {cell} hücresi başarıyla '{value}' olarak güncellendi."
    except Exception as e:
        # Log hatayı (eğer bir logger mevcutsa) ama sadece kısa mesaj döndür.
        return f"Excel dosyasını güncellerken bir hata oluştu: {str(e)}"
